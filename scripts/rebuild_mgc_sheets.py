import sys, math
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

# ── colours (exact v14 dark theme) ─────────────────────────────────────────
C_BG_TITLE   = 'FF0B1120'
C_BG_ALT     = 'FF13161B'
C_BG_DATA    = 'FF0D2E20'
C_BG_PLAIN   = 'FF0D0F12'
C_TEAL   = 'FF00C896'
C_AMBER  = 'FFF5A623'
C_PURPLE = 'FFA855F7'
C_BLUE   = 'FF4A9EFF'
C_DIM    = 'FF7A8494'
C_LIGHT  = 'FFE2E6ED'
C_RED    = 'FFEF4444'

def fill(rgb): return PatternFill('solid', fgColor=rgb)
def font(rgb, bold=False, sz=10): return Font(color=rgb, bold=bold, size=sz, name='Segoe UI')
def align(h='center', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def s(ws, row, col, value=None, bg=C_BG_DATA, fg=C_LIGHT, bold=False, sz=10, h='center', wrap=False):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    c.fill = fill(bg); c.font = font(fg, bold, sz); c.alignment = align(h, wrap=wrap)
    return c

def banner(ws, row, ncols, text, bg=C_BG_TITLE, fg=C_TEAL, sz=10):
    c = ws.cell(row=row, column=1)
    c.value = text; c.fill = fill(bg); c.font = font(fg, True, sz); c.alignment = align('left')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def spacer(ws, row, ncols):
    for col in range(1, ncols+1): ws.cell(row=row, column=col).fill = fill(C_BG_PLAIN)

def clear_sheet(ws, ncols=20):
    for mc in list(ws.merged_cells.ranges): ws.unmerge_cells(str(mc))
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value = None; cell.fill = fill(C_BG_PLAIN)
                cell.font = Font(name='Segoe UI', size=10); cell.alignment = Alignment()
            except AttributeError: pass

# ── MGC contract ─────────────────────────────────────────────────────────
TICK_VAL = 1.0; TICKS_PT = 10; PT_VAL = 10.0; FEE = 1.50
NCOLS = 20

SOP_MAP = {4:'SOP3',5:'SOP3',6:'SOP3',7:'SOP35',8:'SOP4',9:'SOP45',10:'SOP5'}
SOP_TRAIL = {
    'SOP3':  ('3.0pt→1.0pt','4.0pt→2.5pt','5.0pt→4.0pt'),
    'SOP35': ('3.5pt→1.5pt','4.5pt→3.0pt','5.5pt→4.5pt'),
    'SOP4':  ('4.0pt→2.0pt','5.0pt→3.5pt','6.0pt→5.0pt'),
    'SOP45': ('4.5pt→2.5pt','5.5pt→4.0pt','6.5pt→5.5pt'),
    'SOP5':  ('5.0pt→3.0pt','6.0pt→4.5pt','7.0pt→6.0pt'),
}
SOPS = [
    ('SOP3',  'SL4, SL5, SL6', '3.0pt (30tk)','1.0pt', '4.0pt (40tk)','2.5pt', '5.0pt (50tk)','4.0pt', 'T1 fills before trail fires'),
    ('SOP35', 'SL7',            '3.5pt (35tk)','1.5pt', '4.5pt (45tk)','3.0pt', '5.5pt (55tk)','4.5pt', 'Trail fires at T1 (3.5pt = half SL7)'),
    ('SOP4',  'SL8',            '4.0pt (40tk)','2.0pt', '5.0pt (50tk)','3.5pt', '6.0pt (60tk)','5.0pt', 'Trail fires at T1 (4.0pt = half SL8)'),
    ('SOP45', 'SL9',            '4.5pt (45tk)','2.5pt', '5.5pt (55tk)','4.0pt', '6.5pt (65tk)','5.5pt', 'Trail fires at T1 (4.5pt = half SL9)'),
    ('SOP5',  'SL10',           '5.0pt (50tk)','3.0pt', '6.0pt (60tk)','4.5pt', '7.0pt (70tk)','6.0pt', 'Trail fires at T1 (5.0pt = half SL10)'),
]

PROBS = {'Full':0.20,'Trail1':0.10,'Trail2':0.10,'Trail3':0.10,
         'AutoBE':0.10,'HardStop':0.10,'Quick':0.10,'BEbtn':0.20}
WEIGHTS = {4:0.45, 5:0.30, 6:0.15, 7:0.07, 8:0.03}

# old v14 quantities
QTY = {400:{4:10,5:8,6:6,7:5,8:5,9:4,10:4},
       200:{4:5, 5:4,6:3,7:3,8:2,9:2,10:2}}

def compute_atm(sl_pts, total):
    sl_tk = sl_pts * TICKS_PT
    max_risk = total * sl_tk * TICK_VAL
    t1_tk = sl_tk // 2
    t2_tk = int(sl_tk * 0.75)   # floor to whole ticks
    t3_tk = sl_tk
    q1 = math.ceil(total / 2)
    q2 = math.ceil((total - q1) / 2)
    q3 = total - q1 - q2
    fees = total * FEE
    sop = SOP_MAP[sl_pts]
    tr1, tr2, tr3 = SOP_TRAIL[sop]
    return dict(sl_pts=sl_pts, sl_tk=sl_tk, max_risk=max_risk, total=total,
                q1=q1, q2=q2, q3=q3, t1_tk=t1_tk, t2_tk=t2_tk, t3_tk=t3_tk,
                t1_pt=t1_tk/TICKS_PT, t2_pt=t2_tk/TICKS_PT, t3_pt=t3_tk/TICKS_PT,
                be_trig=f'{t1_tk/TICKS_PT:.1f}pt', be_buf='+0.20pt',
                sop=sop, tr1=tr1, tr2=tr2, tr3=tr3, fees=f'${fees:.2f}')

def compute_scenarios(a):
    q1,q2,q3 = a['q1'],a['q2'],a['q3']
    t1,t2,t3 = a['t1_tk'],a['t2_tk'],a['t3_tk']
    total = a['total']; fees = total * FEE
    sl_tk = a['sl_tk']

    full    = (q1*t1 + q2*t2 + q3*t3) * TICK_VAL - fees
    tr1_stp = t1 - 4
    trail1  = (q1*t1 + (q2+q3)*max(tr1_stp,0)) * TICK_VAL - fees
    tr2_stp = t2 - 6
    trail2  = (q1*t1 + q2*t2 + q3*max(tr2_stp,0)) * TICK_VAL - fees
    trail3  = full
    auto_be = (q1*t1 + (q2+q3)*2) * TICK_VAL - fees
    hard    = -(total * sl_tk * TICK_VAL) - fees
    quick   = (q1*t1 + (q2+q3)*(t1//2)) * TICK_VAL - fees
    be_btn  = 0.40*(-fees) + 0.60*hard
    et = (PROBS['Full']*full + PROBS['Trail1']*trail1 + PROBS['Trail2']*trail2 +
          PROBS['Trail3']*trail3 + PROBS['AutoBE']*auto_be +
          PROBS['HardStop']*hard + PROBS['Quick']*quick + PROBS['BEbtn']*be_btn)
    def f(v): return f'+${v:.2f}' if v >= 0 else f'−${abs(v):.2f}'
    return dict(full=f(full), trail1=f(trail1), trail2=f(trail2), trail3=f(trail3),
                auto_be=f(auto_be), hard=f(hard), quick=f(quick), be_btn=f(be_btn), et=f(et),
                et_raw=et)

def compute_expect(a, sc):
    q1,q2,q3 = a['q1'],a['q2'],a['q3']
    t1,t2,t3 = a['t1_tk'],a['t2_tk'],a['t3_tk']
    total = a['total']; fees = total * FEE; sl_tk = a['sl_tk']
    def f(v): return f'+${v:.2f}' if v >= 0 else f'−${abs(v):.2f}'
    full    = (q1*t1 + q2*t2 + q3*t3) * TICK_VAL - fees
    tr1_stp = t1 - 4
    trail1  = (q1*t1 + (q2+q3)*max(tr1_stp,0)) * TICK_VAL - fees
    tr2_stp = t2 - 6
    trail2  = (q1*t1 + q2*t2 + q3*max(tr2_stp,0)) * TICK_VAL - fees
    trail3  = full
    auto_be = (q1*t1 + (q2+q3)*2) * TICK_VAL - fees
    hard    = -(total * sl_tk * TICK_VAL) - fees
    quick   = (q1*t1 + (q2+q3)*(t1//2)) * TICK_VAL - fees
    be_btn  = 0.40*(-fees) + 0.60*hard
    return dict(
        f20=f(PROBS['Full']*full), t1=f(PROBS['Trail1']*trail1),
        t2=f(PROBS['Trail2']*trail2), t3=f(PROBS['Trail3']*trail3),
        abe=f(PROBS['AutoBE']*auto_be), hs=f(PROBS['HardStop']*hard),
        qk=f(PROBS['Quick']*quick), beb=f(PROBS['BEbtn']*be_btn),
        et=f(sc['et_raw']))

def build_sheet(ws, risk_level, title_suffix):
    clear_sheet(ws, NCOLS)
    col_widths = [7,7,8,7,6,6,6, 9,8, 9,8, 9,8, 9,9, 7, 14,14,14, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, 50):
        ws.row_dimensions[r].height = 18
    for r in [1, 5, 15, 25, 36]:
        ws.row_dimensions[r].height = 22 if r == 1 else 30

    slv = [4,5,6,7,8,9,10]
    atms = [compute_atm(sl, QTY[risk_level][sl]) for sl in slv]
    scenarios = [compute_scenarios(a) for a in atms]
    expect = [compute_expect(a, sc) for a, sc in zip(atms, scenarios)]

    # system E
    sys_e = sum(WEIGHTS[a['sl_pts']] * sc['et_raw'] for a, sc in zip(atms, scenarios) if a['sl_pts'] in WEIGHTS)
    sys_e_str = f'+${sys_e:.2f}' if sys_e >= 0 else f'−${abs(sys_e):.2f}'

    # ROW 1 title
    banner(ws, 1, NCOLS,
        f'MGC ATM Grid ${risk_level}  |  3-target SOP trail  |  T1=50%SL  T2=75%SL  T3=100%SL  '
        f'|  BE trigger=T1, +2tk buffer  |  Trail ALL contracts  '
        f'|  $10/pt  $1/tick  fee=$1.50/contract  |  {title_suffix}',
        bg=C_BG_TITLE, fg=C_TEAL, sz=11)
    # ROW 2 subtitle
    banner(ws, 2, NCOLS,
        f'Q1=ceil(total/2)  |  Q2=ceil(rest/2)  |  Q3=remainder  |  '
        f'SOP: SL4-6=SOP3  SL7=SOP35  SL8=SOP4  SL9=SOP45  SL10=SOP5  |  '
        f'System E/trade={sys_e_str}  (weights: SL4=45% SL5=30% SL6=15% SL7=7% SL8=3%)',
        bg=C_BG_ALT, fg=C_DIM, sz=9)
    spacer(ws, 3, NCOLS)

    # ── SECTION 1 ATM CONFIG ────────────────────────────────────────────
    banner(ws, 4, NCOLS, '▶  ATM CONFIG  —  SL4 through SL10', bg=C_BG_TITLE, fg=C_TEAL)
    HDR = [(1,'SL\n(pts)',C_DIM),(2,'SL\n(ticks)',C_DIM),(3,'Max\nRisk',C_LIGHT),(4,'Total\nQty',C_LIGHT),
           (5,'Q1\n(T1)',C_TEAL),(6,'Q2\n(T2)',C_AMBER),(7,'Q3\n(T3)',C_PURPLE),
           (8,'T1\n(pts)',C_TEAL),(9,'T1\n(ticks)',C_TEAL),(10,'T2\n(pts)',C_AMBER),(11,'T2\n(ticks)',C_AMBER),
           (12,'T3\n(pts)',C_PURPLE),(13,'T3\n(ticks)',C_PURPLE),
           (14,'BE\nTrigger',C_BLUE),(15,'BE\nBuffer',C_BLUE),(16,'SOP',C_BLUE),
           (17,'Trail\nStep1',C_DIM),(18,'Trail\nStep2',C_DIM),(19,'Trail\nStep3',C_DIM),(20,'Fees\n(RT)',C_DIM)]
    for col, txt, fg in HDR:
        s(ws, 5, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
    for i, a in enumerate(atms):
        r = 6 + i; bg = C_BG_DATA
        s(ws,r,1, a['sl_pts'],         bg,C_DIM,   sz=10)
        s(ws,r,2, a['sl_tk'],          bg,C_DIM,   sz=10)
        s(ws,r,3, f"${a['max_risk']:.0f}", bg,C_LIGHT, sz=10)
        s(ws,r,4, a['total'],          bg,C_LIGHT, sz=10)
        s(ws,r,5, a['q1'],             bg,C_TEAL,  sz=10,bold=True)
        s(ws,r,6, a['q2'],             bg,C_AMBER, sz=10)
        s(ws,r,7, a['q3'],             bg,C_PURPLE,sz=10)
        s(ws,r,8, f"{a['t1_pt']:.1f}pt",  bg,C_TEAL,  sz=10)
        s(ws,r,9, a['t1_tk'],          bg,C_TEAL,  sz=10)
        s(ws,r,10,f"{a['t2_pt']:.2f}pt", bg,C_AMBER, sz=10)
        s(ws,r,11,a['t2_tk'],          bg,C_AMBER, sz=10)
        s(ws,r,12,f"{a['t3_pt']:.1f}pt",  bg,C_PURPLE,sz=10)
        s(ws,r,13,a['t3_tk'],          bg,C_PURPLE,sz=10)
        s(ws,r,14,a['be_trig'],        bg,C_BLUE,  sz=10)
        s(ws,r,15,a['be_buf'],         bg,C_BLUE,  sz=10)
        s(ws,r,16,a['sop'],            bg,C_BLUE,  sz=10,bold=True)
        s(ws,r,17,a['tr1'],            bg,C_DIM,   sz=9, h='left')
        s(ws,r,18,a['tr2'],            bg,C_DIM,   sz=9, h='left')
        s(ws,r,19,a['tr3'],            bg,C_DIM,   sz=9, h='left')
        s(ws,r,20,a['fees'],           bg,C_DIM,   sz=10)
    spacer(ws, 13, NCOLS)

    # ── SECTION 2 SCENARIO PnL ─────────────────────────────────────────
    banner(ws, 14, NCOLS, '▶  SCENARIO PnL  —  All outcomes per SL  (net after fees)', bg=C_BG_TITLE, fg=C_TEAL)
    SC_HDR = [(1,'SL',C_DIM),(2,'SOP',C_BLUE),(3,'Full\nWin',C_TEAL),(4,'Trail1\nstop',C_TEAL),
              (5,'Trail2\nstop',C_TEAL),(6,'Trail3\nstop',C_TEAL),(7,'Auto BE\n(T1+2tk)',C_AMBER),
              (8,'Hard\nStop',C_RED),(9,'Quick\nBtn',C_AMBER),(10,'BE Btn\n(40/60)',C_PURPLE),(11,'E/trade',C_TEAL)]
    for col, txt, fg in SC_HDR:
        s(ws, 15, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
    for col in range(12, NCOLS+1): s(ws, 15, col, None, bg=C_BG_TITLE)
    for i, (a, sc) in enumerate(zip(atms, scenarios)):
        r = 16 + i; bg = C_BG_DATA
        s(ws,r,1,  f"SL{a['sl_pts']}", bg,C_DIM,   sz=10,bold=True)
        s(ws,r,2,  a['sop'],           bg,C_BLUE,  sz=9)
        s(ws,r,3,  sc['full'],         bg,C_TEAL,  sz=10,bold=True)
        s(ws,r,4,  sc['trail1'],       bg,C_TEAL,  sz=10)
        s(ws,r,5,  sc['trail2'],       bg,C_TEAL,  sz=10)
        s(ws,r,6,  sc['trail3'],       bg,C_TEAL,  sz=10)
        s(ws,r,7,  sc['auto_be'],      bg,C_AMBER, sz=10)
        s(ws,r,8,  sc['hard'],         bg,C_RED,   sz=10)
        s(ws,r,9,  sc['quick'],        bg,C_AMBER, sz=10)
        s(ws,r,10, sc['be_btn'],       bg,C_PURPLE,sz=10)
        s(ws,r,11, sc['et'],           bg,C_TEAL,  sz=10,bold=True)
        for col in range(12, NCOLS+1): s(ws,r,col, None, bg)
    spacer(ws, 23, NCOLS)

    # ── SECTION 3 EXPECTANCY ───────────────────────────────────────────
    banner(ws, 24, NCOLS,
        '▶  EXPECTANCY BREAKDOWN  —  Full=20%  Trail1=10%  Trail2=10%  Trail3=10%  AutoBE=10%  HardStop=10%  Quick=10%  BEbtn=20%',
        bg=C_BG_TITLE, fg=C_TEAL)
    EX_HDR = [(1,'SL',C_DIM),(2,'Full\n20%',C_TEAL),(3,'Trail1\n10%',C_TEAL),(4,'Trail2\n10%',C_TEAL),
              (5,'Trail3\n10%',C_TEAL),(6,'AutoBE\n10%',C_AMBER),(7,'HardStop\n10%',C_RED),
              (8,'Quick\n10%',C_AMBER),(9,'BEbtn\n20%',C_PURPLE),(10,'E/trade',C_TEAL)]
    for col, txt, fg in EX_HDR:
        s(ws, 25, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
    for col in range(11, NCOLS+1): s(ws, 25, col, None, bg=C_BG_TITLE)
    for i, (a, e) in enumerate(zip(atms, expect)):
        r = 26 + i; bg = C_BG_DATA
        s(ws,r,1, f"SL{a['sl_pts']}", bg,C_DIM,   sz=10,bold=True)
        s(ws,r,2, e['f20'],           bg,C_TEAL,  sz=10)
        s(ws,r,3, e['t1'],            bg,C_TEAL,  sz=10)
        s(ws,r,4, e['t2'],            bg,C_TEAL,  sz=10)
        s(ws,r,5, e['t3'],            bg,C_TEAL,  sz=10)
        s(ws,r,6, e['abe'],           bg,C_AMBER, sz=10)
        s(ws,r,7, e['hs'],            bg,C_RED,   sz=10)
        s(ws,r,8, e['qk'],            bg,C_AMBER, sz=10)
        s(ws,r,9, e['beb'],           bg,C_PURPLE,sz=10)
        s(ws,r,10,e['et'],            bg,C_TEAL,  sz=10,bold=True)
        for col in range(11, NCOLS+1): s(ws,r,col, None, bg)
    # system row
    r_sys = 33
    for col in range(1, NCOLS+1): s(ws, r_sys, col, None, bg=C_BG_ALT)
    s(ws, r_sys, 1,  'SYSTEM WEIGHTED E/trade', C_BG_ALT, C_LIGHT, bold=True, sz=10, h='left')
    s(ws, r_sys, 10, sys_e_str,                 C_BG_ALT, C_TEAL,  bold=True, sz=11)
    s(ws, r_sys, 11, 'SL4=45%  SL5=30%  SL6=15%  SL7=7%  SL8=3%', C_BG_ALT, C_DIM, sz=9, h='left')
    ws.merge_cells(start_row=r_sys, start_column=1, end_row=r_sys, end_column=9)
    ws.merge_cells(start_row=r_sys, start_column=11, end_row=r_sys, end_column=NCOLS)
    spacer(ws, 34, NCOLS)

    # ── SECTION 4 SOP REFERENCE ────────────────────────────────────────
    banner(ws, 35, NCOLS,
        '▶  STOP STRATEGY REFERENCE  —  5 SOPs  (trail ALL contracts, freq 2/2/1)  [tick distances 2.5x MES — same SOP names]',
        bg=C_BG_TITLE, fg=C_PURPLE)
    SOP_HDR = [(1,'SOP',C_BLUE),(2,'Used by',C_DIM),
               (3,'Trail trigger\nStep 1',C_TEAL),(4,'Stop after\nStep 1',C_TEAL),
               (5,'Trail trigger\nStep 2',C_AMBER),(6,'Stop after\nStep 2',C_AMBER),
               (7,'Trail trigger\nStep 3',C_PURPLE),(8,'Stop after\nStep 3',C_PURPLE),(9,'Notes',C_DIM)]
    for col, txt, fg in SOP_HDR:
        s(ws, 36, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
    for col in range(10, NCOLS+1): s(ws, 36, col, None, bg=C_BG_TITLE)
    for i, sop in enumerate(SOPS):
        r = 37 + i; bg = C_BG_DATA if i % 2 == 0 else C_BG_ALT
        nm,used,trig1,stop1,trig2,stop2,trig3,stop3,note = sop
        s(ws,r,1,nm,    bg,C_BLUE,  sz=10,bold=True)
        s(ws,r,2,used,  bg,C_LIGHT, sz=9, h='left')
        s(ws,r,3,trig1, bg,C_TEAL,  sz=9)
        s(ws,r,4,stop1, bg,C_TEAL,  sz=9)
        s(ws,r,5,trig2, bg,C_AMBER, sz=9)
        s(ws,r,6,stop2, bg,C_AMBER, sz=9)
        s(ws,r,7,trig3, bg,C_PURPLE,sz=9)
        s(ws,r,8,stop3, bg,C_PURPLE,sz=9)
        s(ws,r,9,note,  bg,C_DIM,   sz=9, h='left')
        for col in range(10, NCOLS+1): s(ws,r,col, None, bg)

# ── Load v18, rebuild both MGC sheets, save as v19 ─────────────────────────
wb = openpyxl.load_workbook(r'C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v18.xlsx')

build_sheet(wb['MGC ATM Grid $400'], 400, 'Qty from MGC v14')
build_sheet(wb['MGC ATM Grid $200'], 200, 'Qty from MGC v14')

out = r'C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v19.xlsx'
wb.save(out)
print(f'Saved: {out}')
print('Sheets:', wb.sheetnames)
