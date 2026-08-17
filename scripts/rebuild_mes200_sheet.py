"""
Rebuild the MES ATM Grid $200 sheet in ATM-Grid-v16.xlsx
with the new 3-target SOP design, matching the $400 sheet structure.
All other sheets are preserved exactly.
"""
import sys
sys.path.insert(0, ".")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SRC = r"C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v16.xlsx"
DST = SRC.replace("ATM-Grid-v16.xlsx", "ATM-Grid-v17.xlsx")

wb = openpyxl.load_workbook(SRC)

# ── remove old $200 sheet, add fresh one at same position ──
sheet_names = wb.sheetnames
pos = sheet_names.index("MES ATM Grid $200")
del wb["MES ATM Grid $200"]
ws = wb.create_sheet("MES ATM Grid $200", pos)

# ── colour palette (matching $400 sheet look) ──
BG_TITLE   = "0D1B2A"   # very dark navy
BG_SUBHEAD = "0D1B2A"
BG_HEADER  = "1A2744"   # mid-navy
BG_DATA    = "0F1923"   # near-black rows
BG_DATA2   = "111F30"   # alternating row
BG_SECTION = "0B1520"   # section divider rows
BG_SYS     = "0D2A1A"   # system row (dark green tint)

FG_TITLE   = "F1C40F"   # yellow
FG_SUB     = "7FB3D3"   # pale blue
FG_HDR     = "93C5FD"   # light blue headers
FG_GREEN   = "4ADE80"
FG_RED     = "F87171"
FG_YELLOW  = "FCD34D"
FG_WHITE   = "E2E8F0"
FG_MUTED   = "64748B"
FG_TEAL    = "34D399"
FG_BLUE    = "60A5FA"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=FG_WHITE, bold=False, size=10, italic=False):
    return Font(name="Segoe UI", color=color, bold=bold, size=size, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(border_style="thin", color="1E2536")
    return Border(left=s, right=s, top=s, bottom=s)

# ── $200 data ──
# Quantities from the previous $200 ATMs, 3-bracket split: Q1=ceil(total/2), Q2=ceil(rest/2), Q3=rest
atms = [
    # sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, max_risk, sop
    (4,  16, 10, 5, 3, 2,  8, 12, 16, 200,  "SOP3"),
    (5,  20,  8, 4, 2, 2, 10, 15, 20, 200,  "SOP3"),
    (6,  24,  7, 4, 2, 1, 12, 18, 24, 210,  "SOP3"),
    (7,  28,  6, 3, 2, 1, 14, 21, 28, 210,  "SOP35"),
    (8,  32,  5, 3, 1, 1, 16, 24, 32, 200,  "SOP4"),
    (9,  36,  5, 3, 1, 1, 18, 27, 36, 225,  "SOP45"),
    (10, 40,  4, 2, 1, 1, 20, 30, 40, 200,  "SOP5"),
]

# scenario PnL (fee = $0.57/contract RT)
# Full Win: all 3 targets hit, no trail interference
# trail1/2/3 stops: Q1 exits T1, Q2+Q3 stopped by trail step 1/2/3
# AutoBE: all stopped at entry+2tk (Q1 at T1, Q2+Q3 at +2tk)
# HardStop: all stopped at SL

def fmt_risk(r):
    return f"${r}"

fee_per = 0.57

def calc_scenarios(sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, sop):
    tpv = 1.25  # $ per tick per contract
    fee = total * fee_per
    t1_val  = t1_tk * tpv
    t2_val  = t2_tk * tpv
    t3_val  = t3_tk * tpv
    sl_val  = sl_tk * tpv

    # trail stop values (stop level in ticks from entry, so profit if exited there)
    sop_stops = {
        "SOP3":  [4, 10, 16],   # step1/2/3 stop ticks
        "SOP35": [6, 12, 18],
        "SOP4":  [8, 14, 20],
        "SOP45": [10, 16, 22],
        "SOP5":  [12, 18, 24],
    }
    stops = sop_stops[sop]

    full_win  = (q1*t1_val + q2*t2_val + q3*t3_val) - fee
    trail1    = (q1*t1_val + q2*stops[0]*tpv + q3*stops[0]*tpv) - fee
    trail2    = (q1*t1_val + q2*t2_val + q3*stops[1]*tpv) - fee
    trail3    = (q1*t1_val + q2*t2_val + q3*t3_val) - fee  # same as full (step3 stop > t3)
    auto_be   = (q1*t1_val + q2*2*tpv + q3*2*tpv) - fee   # Q1 hits T1, rest stopped at +2tk
    hard_stop = -(total * sl_val) - fee
    quick_btn = (q1*t1_val + q2*4*tpv + q3*4*tpv) - fee   # quick exit at ~1pt (4tk)
    be_btn    = 0.40*(-(fee)) + 0.60*(hard_stop)             # 40% scratch, 60% hard stop

    # E/trade probabilities
    probs = {"full":0.20,"trail1":0.10,"trail2":0.10,"trail3":0.10,
             "auto_be":0.10,"hard_stop":0.10,"quick":0.10,"be_btn":0.20}
    e = (probs["full"]*full_win + probs["trail1"]*trail1 + probs["trail2"]*trail2
         + probs["trail3"]*trail3 + probs["auto_be"]*auto_be + probs["hard_stop"]*hard_stop
         + probs["quick"]*quick_btn + probs["be_btn"]*be_btn)

    def f(v):
        sign = "+" if v >= 0 else ""
        return f"{sign}${v:.2f}"

    return {
        "full": f(full_win), "trail1": f(trail1), "trail2": f(trail2), "trail3": f(trail3),
        "auto_be": f(auto_be), "hard_stop": f(hard_stop), "quick": f(quick_btn),
        "be_btn": f(be_btn), "e": f(e), "e_val": e,
        "fee": f"${fee:.2f}"
    }

row = 1

def set_row_bg(ws, r, num_cols, hex_bg):
    for c in range(1, num_cols+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill(hex_bg)

# ── Row 1: Title ──
ws.merge_cells(f"A1:T1")
c = ws["A1"]
c.value = "MES ATM Grid $200  |  Template: 3-5-75-1 heavyT1  |  T1=50%SL  T2=75%SL  T3=100%SL  |  BE trigger=T1, buffer=+2tk (+0.50pt)  |  Trail ALL contracts 8tk/6tk/4tk freq 2/2/1  |  $5/pt  $1.25/tick  fee=$0.57/contract"
c.fill = fill(BG_TITLE); c.font = font(FG_TITLE, bold=True, size=11)
c.alignment = align("center", wrap=True)
ws.row_dimensions[1].height = 28
set_row_bg(ws, 1, 20, BG_TITLE)
row = 2

# ── Row 2: subtitle ──
ws.merge_cells(f"A2:T2")
c = ws["A2"]
c.value = "T1 qty = ceiling(total/2)  |  T2 = ceiling(rest/2)  |  T3 = remainder  |  SOP per SL: SL4-6=SOP3  SL7=SOP35  SL8=SOP4  SL9=SOP45  SL10=SOP5  |  Quantities from previous MES $200 ATMs"
c.fill = fill(BG_TITLE); c.font = font(FG_SUB, size=9, italic=True)
c.alignment = align("center")
ws.row_dimensions[2].height = 16
set_row_bg(ws, 2, 20, BG_TITLE)
row = 3

# ── blank spacer ──
set_row_bg(ws, row, 20, BG_SECTION); ws.row_dimensions[row].height = 6; row += 1

# ── Section: ATM CONFIG ──
ws.merge_cells(f"A{row}:T{row}")
c = ws.cell(row=row, column=1)
c.value = "▶  ATM CONFIG  —  SL4 through SL10"
c.fill = fill(BG_SECTION); c.font = font(FG_YELLOW, bold=True, size=10)
c.alignment = align("left")
ws.row_dimensions[row].height = 20; row += 1

# Config column headers
cfg_headers = ["SL\n(pts)","SL\n(ticks)","Max\nRisk","Total\nQty","T1\nQty","T2\nQty","T3\nQty",
               "T1\n(pts)","T1\n(ticks)","T2\n(pts)","T2\n(ticks)","T3\n(pts)","T3\n(ticks)",
               "BE\nTrigger","BE\nStop","SOP","Trail\nStep1","Trail\nStep2","Trail\nStep3","Fees"]
for ci, h in enumerate(cfg_headers, 1):
    c = ws.cell(row=row, column=ci)
    c.value = h; c.fill = fill(BG_HEADER); c.font = font(FG_HDR, bold=True, size=9)
    c.alignment = align("center", wrap=True); c.border = thin_border()
ws.row_dimensions[row].height = 30; row += 1

sop_trail = {
    "SOP3":  ("3.0pt→1.0pt","4.0pt→2.5pt","5.0pt→4.0pt"),
    "SOP35": ("3.5pt→1.5pt","4.5pt→3.0pt","5.5pt→4.5pt"),
    "SOP4":  ("4.0pt→2.0pt","5.0pt→3.5pt","6.0pt→5.0pt"),
    "SOP45": ("4.5pt→2.5pt","5.5pt→4.0pt","6.5pt→5.5pt"),
    "SOP5":  ("5.0pt→3.0pt","6.0pt→4.5pt","7.0pt→6.0pt"),
}

for i, (sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, max_risk, sop) in enumerate(atms):
    sc = calc_scenarios(sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, sop)
    bg = BG_DATA if i % 2 == 0 else BG_DATA2
    trail = sop_trail[sop]
    row_data = [
        sl_pts, sl_tk, fmt_risk(max_risk), total, q1, q2, q3,
        f"{t1_tk/4:.1f}pt", t1_tk, f"{t2_tk/4:.2f}pt".replace(".00pt",".0pt"), t2_tk,
        f"{t3_tk/4:.1f}pt", t3_tk,
        f"{t1_tk/4:.1f}pt", "+0.50pt", sop, trail[0], trail[1], trail[2], sc["fee"]
    ]
    col_colors = [FG_WHITE,FG_MUTED,FG_RED,FG_YELLOW,FG_BLUE,FG_WHITE,FG_WHITE,
                  FG_GREEN,FG_MUTED,FG_GREEN,FG_MUTED,FG_GREEN,FG_MUTED,
                  FG_YELLOW,FG_TEAL,FG_BLUE,FG_MUTED,FG_MUTED,FG_MUTED,FG_RED]
    for ci, (val, clr) in enumerate(zip(row_data, col_colors), 1):
        c = ws.cell(row=row, column=ci)
        c.value = val; c.fill = fill(bg); c.font = font(clr, bold=(ci in [1,4,16]))
        c.alignment = align(); c.border = thin_border()
    ws.row_dimensions[row].height = 18; row += 1

# ── blank spacer ──
set_row_bg(ws, row, 20, BG_SECTION); ws.row_dimensions[row].height = 6; row += 1

# ── Section: SCENARIO PnL ──
ws.merge_cells(f"A{row}:T{row}")
c = ws.cell(row=row, column=1)
c.value = "▶  SCENARIO PnL  —  All outcomes per SL"
c.fill = fill(BG_SECTION); c.font = font(FG_YELLOW, bold=True, size=10)
c.alignment = align("left"); ws.row_dimensions[row].height = 20; row += 1

pnl_headers = ["SL","SOP","Full\nWin","Trail1\nstop","Trail2\nstop","Trail3\nstop",
               "Auto BE\n(T1+2tk)","Hard\nStop","Quick\nBtn","BE Btn\n(EV 40/60)","E/trade"]
for ci, h in enumerate(pnl_headers, 1):
    c = ws.cell(row=row, column=ci)
    c.value = h; c.fill = fill(BG_HEADER); c.font = font(FG_HDR, bold=True, size=9)
    c.alignment = align("center", wrap=True); c.border = thin_border()
ws.row_dimensions[row].height = 30; row += 1

for i, (sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, max_risk, sop) in enumerate(atms):
    sc = calc_scenarios(sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, sop)
    bg = BG_DATA if i % 2 == 0 else BG_DATA2
    row_data = [f"SL{sl_pts}", sop, sc["full"], sc["trail1"], sc["trail2"], sc["trail3"],
                sc["auto_be"], sc["hard_stop"], sc["quick"], sc["be_btn"], sc["e"]]
    col_colors = [FG_YELLOW, FG_BLUE, FG_GREEN, FG_GREEN, FG_GREEN, FG_GREEN,
                  FG_TEAL, FG_RED, FG_GREEN, FG_RED, FG_WHITE]
    for ci, (val, clr) in enumerate(zip(row_data, col_colors), 1):
        c = ws.cell(row=row, column=ci)
        c.value = val; c.fill = fill(bg); c.font = font(clr, bold=(ci == 11))
        c.alignment = align(); c.border = thin_border()
    ws.row_dimensions[row].height = 18; row += 1

# ── blank spacer ──
set_row_bg(ws, row, 20, BG_SECTION); ws.row_dimensions[row].height = 6; row += 1

# ── Section: EXPECTANCY BREAKDOWN ──
ws.merge_cells(f"A{row}:T{row}")
c = ws.cell(row=row, column=1)
c.value = "▶  EXPECTANCY BREAKDOWN  —  Probs: Full=20% Trail1=10% Trail2=10% Trail3=10% AutoBE=10% HardStop=10% Quick=10% BEbtn=20%"
c.fill = fill(BG_SECTION); c.font = font(FG_YELLOW, bold=True, size=10)
c.alignment = align("left"); ws.row_dimensions[row].height = 20; row += 1

exp_headers = ["SL","Full\n20%","Trail1\n10%","Trail2\n10%","Trail3\n10%",
               "AutoBE\n10%","HardStop\n10%","Quick\n10%","BEbtn\n20%","E/trade"]
for ci, h in enumerate(exp_headers, 1):
    c = ws.cell(row=row, column=ci)
    c.value = h; c.fill = fill(BG_HEADER); c.font = font(FG_HDR, bold=True, size=9)
    c.alignment = align("center", wrap=True); c.border = thin_border()
ws.row_dimensions[row].height = 30; row += 1

sys_e_vals = []
for i, (sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, max_risk, sop) in enumerate(atms):
    sc = calc_scenarios(sl_pts, sl_tk, total, q1, q2, q3, t1_tk, t2_tk, t3_tk, sop)
    bg = BG_DATA if i % 2 == 0 else BG_DATA2
    probs = [0.20,0.10,0.10,0.10,0.10,0.10,0.10,0.20]
    sc_vals_raw = []
    tpv=1.25; fee=total*fee_per
    stops_map={"SOP3":[4,10,16],"SOP35":[6,12,18],"SOP4":[8,14,20],"SOP45":[10,16,22],"SOP5":[12,18,24]}
    st=stops_map[sop]
    full_win  = q1*t1_tk*tpv + q2*t2_tk*tpv + q3*t3_tk*tpv - fee
    trail1    = q1*t1_tk*tpv + q2*st[0]*tpv + q3*st[0]*tpv - fee
    trail2    = q1*t1_tk*tpv + q2*t2_tk*tpv + q3*st[1]*tpv - fee
    trail3    = full_win
    auto_be   = q1*t1_tk*tpv + q2*2*tpv + q3*2*tpv - fee
    hard_stop = -(total*sl_tk*tpv) - fee
    quick_btn = q1*t1_tk*tpv + q2*4*tpv + q3*4*tpv - fee
    be_btn    = 0.40*(-fee) + 0.60*hard_stop
    sc_vals_raw = [full_win, trail1, trail2, trail3, auto_be, hard_stop, quick_btn, be_btn]
    weighted = [p*v for p,v in zip(probs, sc_vals_raw)]
    e_val = sum(weighted)
    sys_e_vals.append((sl_pts, e_val))

    def f2(v):
        sign = "+" if v >= 0 else ""
        return f"{sign}${v:.2f}"

    exp_row = [f"SL{sl_pts}"] + [f2(p*v) for p,v in zip(probs,sc_vals_raw)] + [f2(e_val)]
    col_colors = [FG_YELLOW] + [FG_GREEN if v>=0 else FG_RED for v in sc_vals_raw] + [FG_WHITE]
    for ci, (val, clr) in enumerate(zip(exp_row, col_colors), 1):
        c = ws.cell(row=row, column=ci)
        c.value = val; c.fill = fill(bg); c.font = font(clr, bold=(ci==10))
        c.alignment = align(); c.border = thin_border()
    ws.row_dimensions[row].height = 18; row += 1

# Weighted system row (SL4=45% SL5=30% SL6=15% SL7=7% SL8=3%)
weights = {4:0.45, 5:0.30, 6:0.15, 7:0.07, 8:0.03, 9:0.00, 10:0.00}
sys_e = sum(w * e for sl, e in sys_e_vals for sl2, w in weights.items() if sl == sl2)
ws.merge_cells(f"A{row}:I{row}")
c = ws.cell(row=row, column=1)
c.value = "SYSTEM WEIGHTED E/trade"
c.fill = fill(BG_SYS); c.font = font(FG_GREEN, bold=True, size=10); c.alignment = align("center")
c2 = ws.cell(row=row, column=10)
c2.value = f"+${sys_e:.2f}"; c2.fill = fill(BG_SYS); c2.font = font(FG_GREEN, bold=True, size=11)
c2.alignment = align()
ws.merge_cells(f"K{row}:T{row}")
c3 = ws.cell(row=row, column=11)
c3.value = "SL4=45% SL5=30% SL6=15% SL7=7% SL8=3%"
c3.fill = fill(BG_SYS); c3.font = font(FG_MUTED, size=9, italic=True); c3.alignment = align("left")
ws.row_dimensions[row].height = 22; row += 1

# ── blank spacer ──
set_row_bg(ws, row, 20, BG_SECTION); ws.row_dimensions[row].height = 6; row += 1

# ── Section: SOP REFERENCE ──
ws.merge_cells(f"A{row}:T{row}")
c = ws.cell(row=row, column=1)
c.value = "▶  STOP STRATEGY REFERENCE  —  5 SOPs  (trail ALL contracts, 8tk/6tk/4tk, freq 2/2/1)"
c.fill = fill(BG_SECTION); c.font = font(FG_YELLOW, bold=True, size=10)
c.alignment = align("left"); ws.row_dimensions[row].height = 20; row += 1

sop_headers = ["SOP","Used by","Trail trigger\nStep 1 (8tk)","Stop after\nStep 1",
               "Trail trigger\nStep 2 (6tk)","Stop after\nStep 2",
               "Trail trigger\nStep 3 (4tk)","Stop after\nStep 3","Notes"]
for ci, h in enumerate(sop_headers, 1):
    c = ws.cell(row=row, column=ci)
    c.value = h; c.fill = fill(BG_HEADER); c.font = font(FG_HDR, bold=True, size=9)
    c.alignment = align("center", wrap=True); c.border = thin_border()
ws.row_dimensions[row].height = 30; row += 1

sop_data = [
    ("SOP3","SL4, SL5, SL6","3.0pt (12tk)","1.0pt","4.0pt (16tk)","2.5pt","5.0pt (20tk)","4.0pt","SL4/5/6 — T1 fills before trail fires"),
    ("SOP35","SL7","3.5pt (14tk)","1.5pt","4.5pt (18tk)","3.0pt","5.5pt (22tk)","4.5pt","SL7 — trail fires at T1 level (half SL=3.5pt)"),
    ("SOP4","SL8","4.0pt (16tk)","2.0pt","5.0pt (20tk)","3.5pt","6.0pt (24tk)","5.0pt","SL8 — trail fires at T1 level (half SL=4pt)"),
    ("SOP45","SL9","4.5pt (18tk)","2.5pt","5.5pt (22tk)","4.0pt","6.5pt (26tk)","5.5pt","SL9 — trail fires at T1 level (half SL=4.5pt)"),
    ("SOP5","SL10","5.0pt (20tk)","3.0pt","6.0pt (24tk)","4.5pt","7.0pt (28tk)","6.0pt","SL10 — trail fires at T1 level (half SL=5pt)"),
]
for i, srow in enumerate(sop_data):
    bg = BG_DATA if i % 2 == 0 else BG_DATA2
    for ci, val in enumerate(srow, 1):
        c = ws.cell(row=row, column=ci)
        c.value = val; c.fill = fill(bg)
        c.font = font(FG_BLUE if ci==1 else (FG_MUTED if ci==2 else (FG_YELLOW if ci%2==1 else FG_GREEN)))
        c.alignment = align("center" if ci < 9 else "left"); c.border = thin_border()
    ws.row_dimensions[row].height = 18; row += 1

# ── column widths ──
col_widths = [6,6,8,6,5,5,5, 7,6,8,6,7,6, 8,8,7,13,13,13,7]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(DST)
print(f"Saved: {DST}")
print(f"MES $200 sheet rebuilt with 3-target SOP design. System E/trade = +${sys_e:.2f}")
