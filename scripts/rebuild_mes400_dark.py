import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

# ── colour palette (exact v14 values) ──────────────────────────────────────
C_BG_TITLE   = 'FF0B1120'   # near-black navy — title, section banners, col headers
C_BG_ALT     = 'FF13161B'   # dark grey — subtitle, alt data rows
C_BG_DATA    = 'FF0D2E20'   # dark green-tinted — primary data rows
C_BG_DIVIDER = 'FF0D0F12'   # almost-black — divider columns
C_BG_PLAIN   = 'FF0D0F12'   # empty rows / spacer

C_TEAL   = 'FF00C896'   # T1, NET, section header text, positive
C_AMBER  = 'FFF5A623'   # T2 columns
C_PURPLE = 'FFA855F7'   # T3 columns / section banner variants
C_BLUE   = 'FF4A9EFF'   # SOP, NT ref, split info
C_DIM    = 'FF7A8494'   # SL pts, fees, muted labels
C_LIGHT  = 'FFE2E6ED'   # qty, gross, neutral data
C_WHITE  = 'FFffffff'   # rarely used
C_RED    = 'FFEF4444'   # hard stop / negative
C_GREEN  = 'FF22C55E'   # positive PnL cells

def fill(rgb): return PatternFill('solid', fgColor=rgb)
def font(rgb, bold=False, sz=10): return Font(color=rgb, bold=bold, size=sz, name='Segoe UI')
def align(h='center', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style(ws, row, col, value=None, bg=C_BG_DATA, fg=C_LIGHT, bold=False, sz=10,
          h='center', wrap=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.fill = fill(bg)
    c.font = font(fg, bold=bold, sz=sz)
    c.alignment = align(h, wrap=wrap)
    return c

def banner(ws, row, ncols, text, bg=C_BG_TITLE, fg=C_TEAL, sz=10):
    c = ws.cell(row=row, column=1)
    c.value = text
    c.fill = fill(bg)
    c.font = font(fg, bold=True, sz=sz)
    c.alignment = align('left')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def spacer(ws, row, ncols):
    for col in range(1, ncols+1):
        ws.cell(row=row, column=col).fill = fill(C_BG_PLAIN)

# ── ATM data ────────────────────────────────────────────────────────────────
# SL(pts), SL(tk), MaxRisk, Total, Q1, Q2, Q3, T1pts, T1tk, T2pts, T2tk, T3pts, T3tk, BEtrigger, BEbuf, SOP, Trail1, Trail2, Trail3, Fees
ATMS = [
    (4,  16, '$400', 20, 10,5,5,  '2.0pt',8,  '3.0pt',12,  '4.0pt',16,  '2.0pt','+0.50pt','SOP3',  '3.0pt→1.0pt','4.0pt→2.5pt','5.0pt→4.0pt', '$11.40'),
    (5,  20, '$400', 16,  8,4,4,  '2.5pt',10, '3.75pt',15, '5.0pt',20,  '2.5pt','+0.50pt','SOP3',  '3.0pt→1.0pt','4.0pt→2.5pt','5.0pt→4.0pt', '$9.12'),
    (6,  24, '$390', 13,  7,3,3,  '3.0pt',12, '4.5pt',18,  '6.0pt',24,  '3.0pt','+0.50pt','SOP3',  '3.0pt→1.0pt','4.0pt→2.5pt','5.0pt→4.0pt', '$7.41'),
    (7,  28, '$385', 11,  6,3,2,  '3.5pt',14, '5.25pt',21, '7.0pt',28,  '3.5pt','+0.50pt','SOP35', '3.5pt→1.5pt','4.5pt→3.0pt','5.5pt→4.5pt', '$6.27'),
    (8,  32, '$400', 10,  5,3,2,  '4.0pt',16, '6.0pt',24,  '8.0pt',32,  '4.0pt','+0.50pt','SOP4',  '4.0pt→2.0pt','5.0pt→3.5pt','6.0pt→5.0pt', '$5.70'),
    (9,  36, '$360',  8,  4,2,2,  '4.5pt',18, '6.75pt',27, '9.0pt',36,  '4.5pt','+0.50pt','SOP45', '4.5pt→2.5pt','5.5pt→4.0pt','6.5pt→5.5pt', '$4.56'),
    (10, 40, '$400',  8,  4,2,2,  '5.0pt',20, '7.5pt',30,  '10.0pt',40, '5.0pt','+0.50pt','SOP5',  '5.0pt→3.0pt','6.0pt→4.5pt','7.0pt→6.0pt', '$4.56'),
]

# ── Scenario PnL  ────────────────────────────────────────────────────────────
# All net values (gross - fees).  fee = qty * $0.57
# Full win = Q1*T1 + Q2*T2 + Q3*T3   Trail1 = all@T1+2tk   Trail2 = Q2/Q3 continue   Trail3 = Q3 continues
# AutoBE = Q1*T1 + (Q2+Q3)*BE(+2tk)   HardStop = -SL*total*1.25 - fees
# Quick = Q1@T1 + rest@~T1   BEbtn = 40%*(−fees) + 60%*(HardStop)
# Using stored values from v17 (already verified)
SCENARIOS = [
    # SL, SOP, FullWin, Trail1, Trail2, Trail3, AutoBE, HardStop, Quick, BEbtn, Etrade
    ('SL4', 'SOP3',  '+$263.60','+$188.60','+$263.60','+$263.60','+$113.60','−$411.40','+$138.60','−$251.40','+$58.10'),
    ('SL5', 'SOP3',  '+$265.88','+$130.88','+$215.88','+$265.88','+$110.88','−$409.12','+$110.88','−$249.12','+$45.88'),
    ('SL6', 'SOP3',  '+$255.09','+$127.59','+$172.59','+$225.09','+$112.59','−$397.41', '+$87.59','−$241.41','+$35.54'),
    ('SL7', 'SOP35', '+$247.48','+$136.23','+$173.73','+$222.48','+$111.23','−$391.27', '+$73.73','−$237.27','+$34.66'),
    ('SL8', 'SOP4',  '+$264.30','+$144.30','+$181.80','+$234.30','+$106.80','−$405.70', '+$69.30','−$245.70','+$36.80'),
    ('SL9', 'SOP45', '+$242.94','+$135.44','+$165.44','+$195.44', '+$95.44','−$364.56', '+$55.44','−$220.56','+$32.74'),
    ('SL10','SOP5',  '+$270.44','+$155.44','+$185.44','+$215.44','+$105.44','−$404.56', '+$55.44','−$244.56','+$36.44'),
]

# ── Expectancy breakdown ──────────────────────────────────────────────────
# Full=20% Trail1=10% Trail2=10% Trail3=10% AutoBE=10% HardStop=10% Quick=10% BEbtn=20%
EXPECT = [
    ('SL4', '+$52.72','+$18.86','+$26.36','+$26.36','+$11.36','−$41.14','+$13.86','−$50.28','+$58.10'),
    ('SL5', '+$53.18','+$13.09','+$21.59','+$26.59','+$11.09','−$40.91','+$11.09','−$49.82','+$45.88'),
    ('SL6', '+$51.02','+$12.76','+$17.26','+$22.51','+$11.26','−$39.74', '+$8.76','−$48.28','+$35.54'),
    ('SL7', '+$49.50','+$13.62','+$17.37','+$22.25','+$11.12','−$39.13', '+$7.37','−$47.45','+$34.66'),
    ('SL8', '+$52.86','+$14.43','+$18.18','+$23.43','+$10.68','−$40.57', '+$6.93','−$49.14','+$36.80'),
    ('SL9', '+$48.59','+$13.54','+$16.54','+$19.54', '+$9.54','−$36.46', '+$5.54','−$44.11','+$32.74'),
    ('SL10','+$54.09','+$15.54','+$18.54','+$21.54','+$10.54','−$40.46', '+$5.54','−$48.91','+$36.44'),
]

SOPS = [
    ('SOP3',  'SL4, SL5, SL6', '3.0pt (12tk)','1.0pt', '4.0pt (16tk)','2.5pt', '5.0pt (20tk)','4.0pt', 'T1 fills before trail fires'),
    ('SOP35', 'SL7',            '3.5pt (14tk)','1.5pt', '4.5pt (18tk)','3.0pt', '5.5pt (22tk)','4.5pt', 'Trail fires at T1 (3.5pt = half SL)'),
    ('SOP4',  'SL8',            '4.0pt (16tk)','2.0pt', '5.0pt (20tk)','3.5pt', '6.0pt (24tk)','5.0pt', 'Trail fires at T1 (4.0pt = half SL)'),
    ('SOP45', 'SL9',            '4.5pt (18tk)','2.5pt', '5.5pt (22tk)','4.0pt', '6.5pt (26tk)','5.5pt', 'Trail fires at T1 (4.5pt = half SL)'),
    ('SOP5',  'SL10',           '5.0pt (20tk)','3.0pt', '6.0pt (24tk)','4.5pt', '7.0pt (28tk)','6.0pt', 'Trail fires at T1 (5.0pt = half SL)'),
]

NCOLS = 20

# ── Load workbook ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(r'C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v17.xlsx')
ws = wb['MES ATM Grid $400']

# clear merges first, then clear cells
for mc in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(mc))

for row in ws.iter_rows():
    for cell in row:
        try:
            cell.value = None
            cell.fill = fill('FF0D0F12')
            cell.font = Font(name='Segoe UI', size=10)
            cell.alignment = Alignment()
        except AttributeError:
            pass

# ── Column widths ──────────────────────────────────────────────────────────
col_widths = [7,7,8,7,6,6,6, 8,7, 9,7, 9,7, 9,9, 7, 13,13,13, 8]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Row heights ───────────────────────────────────────────────────────────
for r in range(1, 80):
    ws.row_dimensions[r].height = 18
ws.row_dimensions[1].height = 22
ws.row_dimensions[5].height = 30
ws.row_dimensions[15].height = 30
ws.row_dimensions[25].height = 30
ws.row_dimensions[36].height = 30

# ═══════════════════════════════════════════════════════════════════════════
# ROW 1 — title
# ═══════════════════════════════════════════════════════════════════════════
banner(ws, 1, NCOLS,
       'MES ATM Grid $400  |  3-target SOP trail  |  T1=50%SL  T2=75%SL  T3=100%SL  '
       '|  BE trigger=T1, +2tk buffer  |  Trail ALL contracts 8tk/6tk/4tk freq 2/2/1  '
       '|  $5/pt  $1.25/tick  fee=$0.57/contract',
       bg=C_BG_TITLE, fg=C_TEAL, sz=11)

# ROW 2 — subtitle
banner(ws, 2, NCOLS,
       'Q1=ceil(total/2)  |  Q2=ceil(rest/2)  |  Q3=remainder  |  '
       'SOP: SL4-6=SOP3  SL7=SOP35  SL8=SOP4  SL9=SOP45  SL10=SOP5  |  System E/trade=+$48.77  (weights: SL4=45% SL5=30% SL6=15% SL7=7% SL8=3%)',
       bg=C_BG_ALT, fg=C_DIM, sz=9)

# ROW 3 — spacer
spacer(ws, 3, NCOLS)

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 — ATM CONFIG
# ═══════════════════════════════════════════════════════════════════════════
banner(ws, 4, NCOLS, '▶  ATM CONFIG  —  SL4 through SL10', bg=C_BG_TITLE, fg=C_TEAL)

# headers row 5
HDR_COL = [
    (1,  'SL\n(pts)',   C_DIM),
    (2,  'SL\n(ticks)', C_DIM),
    (3,  'Max\nRisk',   C_LIGHT),
    (4,  'Total\nQty',  C_LIGHT),
    (5,  'Q1\n(T1)',    C_TEAL),
    (6,  'Q2\n(T2)',    C_AMBER),
    (7,  'Q3\n(T3)',    C_PURPLE),
    (8,  'T1\n(pts)',   C_TEAL),
    (9,  'T1\n(ticks)', C_TEAL),
    (10, 'T2\n(pts)',   C_AMBER),
    (11, 'T2\n(ticks)', C_AMBER),
    (12, 'T3\n(pts)',   C_PURPLE),
    (13, 'T3\n(ticks)', C_PURPLE),
    (14, 'BE\nTrigger', C_BLUE),
    (15, 'BE\nBuffer',  C_BLUE),
    (16, 'SOP',         C_BLUE),
    (17, 'Trail\nStep1',C_DIM),
    (18, 'Trail\nStep2',C_DIM),
    (19, 'Trail\nStep3',C_DIM),
    (20, 'Fees\n(RT)',   C_DIM),
]
for col, txt, fg in HDR_COL:
    style(ws, 5, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)

# data rows 6–12
for i, a in enumerate(ATMS):
    r = 6 + i
    sl_pts,sl_tk,maxr,tot,q1,q2,q3,t1p,t1t,t2p,t2t,t3p,t3t,be_trig,be_buf,sop,tr1,tr2,tr3,fees = a
    bg = C_BG_DATA
    style(ws, r, 1,  sl_pts, bg, C_DIM,    sz=10)
    style(ws, r, 2,  sl_tk,  bg, C_DIM,    sz=10)
    style(ws, r, 3,  maxr,   bg, C_LIGHT,  sz=10)
    style(ws, r, 4,  tot,    bg, C_LIGHT,  sz=10)
    style(ws, r, 5,  q1,     bg, C_TEAL,   sz=10, bold=True)
    style(ws, r, 6,  q2,     bg, C_AMBER,  sz=10)
    style(ws, r, 7,  q3,     bg, C_PURPLE, sz=10)
    style(ws, r, 8,  t1p,    bg, C_TEAL,   sz=10)
    style(ws, r, 9,  t1t,    bg, C_TEAL,   sz=10)
    style(ws, r, 10, t2p,    bg, C_AMBER,  sz=10)
    style(ws, r, 11, t2t,    bg, C_AMBER,  sz=10)
    style(ws, r, 12, t3p,    bg, C_PURPLE, sz=10)
    style(ws, r, 13, t3t,    bg, C_PURPLE, sz=10)
    style(ws, r, 14, be_trig,bg, C_BLUE,   sz=10)
    style(ws, r, 15, be_buf, bg, C_BLUE,   sz=10)
    style(ws, r, 16, sop,    bg, C_BLUE,   sz=10, bold=True)
    style(ws, r, 17, tr1,    bg, C_DIM,    sz=9,  h='left')
    style(ws, r, 18, tr2,    bg, C_DIM,    sz=9,  h='left')
    style(ws, r, 19, tr3,    bg, C_DIM,    sz=9,  h='left')
    style(ws, r, 20, fees,   bg, C_DIM,    sz=10)

spacer(ws, 13, NCOLS)

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — SCENARIO PnL
# ═══════════════════════════════════════════════════════════════════════════
banner(ws, 14, NCOLS, '▶  SCENARIO PnL  —  All outcomes per SL  (net after fees)', bg=C_BG_TITLE, fg=C_TEAL)

SC_HDR = [
    (1,  'SL',              C_DIM),
    (2,  'SOP',             C_BLUE),
    (3,  'Full\nWin',       C_TEAL),
    (4,  'Trail1\nstop',    C_TEAL),
    (5,  'Trail2\nstop',    C_TEAL),
    (6,  'Trail3\nstop',    C_TEAL),
    (7,  'Auto BE\n(T1+2tk)',C_AMBER),
    (8,  'Hard\nStop',      C_RED),
    (9,  'Quick\nBtn',      C_AMBER),
    (10, 'BE Btn\n(40/60)', C_PURPLE),
    (11, 'E/trade',         C_TEAL),
]
for col, txt, fg in SC_HDR:
    style(ws, 15, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
# blank remaining cols in header row
for col in range(12, NCOLS+1):
    style(ws, 15, col, None, bg=C_BG_TITLE)

for i, s in enumerate(SCENARIOS):
    r = 16 + i
    bg = C_BG_DATA
    sl, sop, fw, tr1, tr2, tr3, abe, hs, qk, beb, et = s
    style(ws, r, 1,  sl,  bg, C_DIM,    sz=10, bold=True)
    style(ws, r, 2,  sop, bg, C_BLUE,   sz=9)
    style(ws, r, 3,  fw,  bg, C_TEAL,   sz=10, bold=True)
    style(ws, r, 4,  tr1, bg, C_TEAL,   sz=10)
    style(ws, r, 5,  tr2, bg, C_TEAL,   sz=10)
    style(ws, r, 6,  tr3, bg, C_TEAL,   sz=10)
    style(ws, r, 7,  abe, bg, C_AMBER,  sz=10)
    style(ws, r, 8,  hs,  bg, C_RED,    sz=10)
    style(ws, r, 9,  qk,  bg, C_AMBER,  sz=10)
    style(ws, r, 10, beb, bg, C_PURPLE, sz=10)
    style(ws, r, 11, et,  bg, C_TEAL,   sz=10, bold=True)
    for col in range(12, NCOLS+1):
        style(ws, r, col, None, bg)

spacer(ws, 23, NCOLS)

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — EXPECTANCY BREAKDOWN
# ═══════════════════════════════════════════════════════════════════════════
banner(ws, 24, NCOLS,
       '▶  EXPECTANCY BREAKDOWN  —  Full=20%  Trail1=10%  Trail2=10%  Trail3=10%  AutoBE=10%  HardStop=10%  Quick=10%  BEbtn=20%',
       bg=C_BG_TITLE, fg=C_TEAL)

EX_HDR = [
    (1,  'SL',          C_DIM),
    (2,  'Full\n20%',   C_TEAL),
    (3,  'Trail1\n10%', C_TEAL),
    (4,  'Trail2\n10%', C_TEAL),
    (5,  'Trail3\n10%', C_TEAL),
    (6,  'AutoBE\n10%', C_AMBER),
    (7,  'HardStop\n10%',C_RED),
    (8,  'Quick\n10%',  C_AMBER),
    (9,  'BEbtn\n20%',  C_PURPLE),
    (10, 'E/trade',     C_TEAL),
]
for col, txt, fg in EX_HDR:
    style(ws, 25, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
for col in range(11, NCOLS+1):
    style(ws, 25, col, None, bg=C_BG_TITLE)

for i, e in enumerate(EXPECT):
    r = 26 + i
    bg = C_BG_DATA
    sl, f20, t1, t2, t3, abe, hs, qk, beb, et = e
    style(ws, r, 1,  sl,  bg, C_DIM,    sz=10, bold=True)
    style(ws, r, 2,  f20, bg, C_TEAL,   sz=10)
    style(ws, r, 3,  t1,  bg, C_TEAL,   sz=10)
    style(ws, r, 4,  t2,  bg, C_TEAL,   sz=10)
    style(ws, r, 5,  t3,  bg, C_TEAL,   sz=10)
    style(ws, r, 6,  abe, bg, C_AMBER,  sz=10)
    style(ws, r, 7,  hs,  bg, C_RED,    sz=10)
    style(ws, r, 8,  qk,  bg, C_AMBER,  sz=10)
    style(ws, r, 9,  beb, bg, C_PURPLE, sz=10)
    style(ws, r, 10, et,  bg, C_TEAL,   sz=10, bold=True)
    for col in range(11, NCOLS+1):
        style(ws, r, col, None, bg)

# System weighted summary row
r_sys = 33
for col in range(1, NCOLS+1):
    style(ws, r_sys, col, None, bg=C_BG_ALT)
style(ws, r_sys, 1,  'SYSTEM WEIGHTED E/trade', C_BG_ALT, C_LIGHT, bold=True, sz=10, h='left')
style(ws, r_sys, 10, '+$48.77',                 C_BG_ALT, C_TEAL,  bold=True, sz=11)
style(ws, r_sys, 11, 'SL4=45%  SL5=30%  SL6=15%  SL7=7%  SL8=3%', C_BG_ALT, C_DIM, sz=9, h='left')
ws.merge_cells(start_row=r_sys, start_column=1, end_row=r_sys, end_column=9)
ws.merge_cells(start_row=r_sys, start_column=11, end_row=r_sys, end_column=NCOLS)

spacer(ws, 34, NCOLS)

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 — SOP REFERENCE
# ═══════════════════════════════════════════════════════════════════════════
banner(ws, 35, NCOLS,
       '▶  STOP STRATEGY REFERENCE  —  5 SOPs  (trail ALL contracts, 8tk/6tk/4tk, freq 2/2/1)',
       bg=C_BG_TITLE, fg=C_PURPLE)

SOP_HDR = [
    (1, 'SOP',              C_BLUE),
    (2, 'Used by',          C_DIM),
    (3, 'Trail trigger\nStep 1 (8tk)', C_TEAL),
    (4, 'Stop after\nStep 1',          C_TEAL),
    (5, 'Trail trigger\nStep 2 (6tk)', C_AMBER),
    (6, 'Stop after\nStep 2',          C_AMBER),
    (7, 'Trail trigger\nStep 3 (4tk)', C_PURPLE),
    (8, 'Stop after\nStep 3',          C_PURPLE),
    (9, 'Notes',            C_DIM),
]
for col, txt, fg in SOP_HDR:
    style(ws, 36, col, txt, bg=C_BG_TITLE, fg=fg, bold=True, sz=9, wrap=True)
for col in range(10, NCOLS+1):
    style(ws, 36, col, None, bg=C_BG_TITLE)

for i, s in enumerate(SOPS):
    r = 37 + i
    bg = C_BG_DATA if i % 2 == 0 else C_BG_ALT
    sop, used, trig1, stop1, trig2, stop2, trig3, stop3, note = s
    style(ws, r, 1, sop,   bg, C_BLUE,   sz=10, bold=True)
    style(ws, r, 2, used,  bg, C_LIGHT,  sz=9,  h='left')
    style(ws, r, 3, trig1, bg, C_TEAL,   sz=9)
    style(ws, r, 4, stop1, bg, C_TEAL,   sz=9)
    style(ws, r, 5, trig2, bg, C_AMBER,  sz=9)
    style(ws, r, 6, stop2, bg, C_AMBER,  sz=9)
    style(ws, r, 7, trig3, bg, C_PURPLE, sz=9)
    style(ws, r, 8, stop3, bg, C_PURPLE, sz=9)
    style(ws, r, 9, note,  bg, C_DIM,    sz=9,  h='left')
    for col in range(10, NCOLS+1):
        style(ws, r, col, None, bg)

# ── Save ───────────────────────────────────────────────────────────────────
out = r'C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v18.xlsx'
wb.save(out)
print(f'Saved: {out}')
