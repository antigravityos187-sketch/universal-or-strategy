"""Build ATM-Grid-v16.xlsx
Copies all tabs from v14 unchanged, replaces only 'MES ATM Grid $400'.
New config: 3-target (T1=50%SL, T2=75%SL, T3=100%SL), heavy T1 split,
            BE trigger=T1, BE buffer=+2ticks, SOP per SL (half-SL trail trigger).
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math, copy

SRC  = r'C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v14.xlsx'
DEST = r'C:\Users\Mohammed Khalid\AppData\Local\Temp\bob-artifacts\ATM-Grid-v16.xlsx'

TAB_TO_REPLACE = 'MES ATM Grid $400'

# ── ATM DATA ─────────────────────────────────────────────────────
tick_val = 1.25
fee_rt   = 0.57

def max_qty(sl):   return int(400 / (sl * 5))
def pts(v, q):     return round(v * 4 * tick_val * q, 2)
def fees(q):       return round(fee_rt * q, 2)
def tgt(sl, pct):  return round(sl * pct, 2)
def ticks(pt):     return int(round(pt * 4))

def split3(total):
    q1 = (total + 1) // 2
    rest = total - q1
    q2 = (rest + 1) // 2
    q3 = rest - q2
    return q1, q2, q3

sop_map = {4:'SOP3',5:'SOP3',6:'SOP3',7:'SOP35',8:'SOP4',9:'SOP45',10:'SOP5'}
trail_sops = {
    'SOP3':  [(3.0,2.0),(4.0,1.5),(5.0,1.0)],
    'SOP35': [(3.5,2.0),(4.5,1.5),(5.5,1.0)],
    'SOP4':  [(4.0,2.0),(5.0,1.5),(6.0,1.0)],
    'SOP45': [(4.5,2.0),(5.5,1.5),(6.5,1.0)],
    'SOP5':  [(5.0,2.0),(6.0,1.5),(7.0,1.0)],
}

probs = {'full':0.20,'trail1':0.10,'trail2':0.10,'trail3':0.10,
         'auto_be':0.10,'stop':0.10,'quick':0.10,'be_btn':0.20}
weights = {4:0.45,5:0.30,6:0.15,7:0.07,8:0.03,9:0.0,10:0.0}

def trail_pnl(legs, trigger, stop, fee_val):
    return round(sum(pts(tv,q) if tv<=trigger else pts(stop,q) for q,tv in legs) - fee_val, 2)

def calc(sl):
    total = max_qty(sl)
    q1,q2,q3 = split3(total)
    fee = fees(total)
    t1v=tgt(sl,0.5); t2v=tgt(sl,0.75); t3v=tgt(sl,1.0)
    sop=sop_map[sl]; steps=trail_sops[sop]
    stops=[round(s[0]-s[1],2) for s in steps]
    legs=[(q1,t1v),(q2,t2v),(q3,t3v)]
    full  = pts(t1v,q1)+pts(t2v,q2)+pts(t3v,q3)-fee
    tr    = [trail_pnl(legs,steps[i][0],stops[i],fee) for i in range(3)]
    auto_be = pts(t1v,q1)+pts(0.50,q2)+pts(0.50,q3)-fee
    hard  = -pts(sl,total)-fee
    quick = pts(1.0,q1)+pts(2.0,q2)+pts(2.0,q3)-fee
    be_btn= 0.40*(0-fee)+0.60*hard
    e = (probs['full']*full+probs['trail1']*tr[0]+probs['trail2']*tr[1]+
         probs['trail3']*tr[2]+probs['auto_be']*auto_be+probs['stop']*hard+
         probs['quick']*quick+probs['be_btn']*be_btn)
    return dict(sl=sl,total=total,q1=q1,q2=q2,q3=q3,fee=fee,
                t1v=t1v,t2v=t2v,t3v=t3v,sop=sop,steps=steps,stops=stops,
                full=full,tr1=tr[0],tr2=tr[1],tr3=tr[2],
                auto_be=auto_be,hard=hard,quick=quick,be_btn=round(be_btn,2),
                e=round(e,2))

data = [calc(sl) for sl in range(4,11)]

# ── STYLES ───────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1F3864')
HDR2_FILL = PatternFill('solid', fgColor='2E75B6')
HDR3_FILL = PatternFill('solid', fgColor='305496')
ROW_FILL  = PatternFill('solid', fgColor='D9E1F2')
ROW2_FILL = PatternFill('solid', fgColor='EEF2F8')
GRN_FILL  = PatternFill('solid', fgColor='E2EFDA')
RED_FILL  = PatternFill('solid', fgColor='FCE4D6')
YLW_FILL  = PatternFill('solid', fgColor='FFF2CC')
SOP_FILL  = PatternFill('solid', fgColor='F2F2F2')

WHT  = Font(color='FFFFFF', bold=True, size=10)
BLD  = Font(bold=True, size=10)
NRM  = Font(size=10)
SML  = Font(size=9, italic=True)

def bd(v='thin'):
    s = Side(style=v)
    return Border(left=s,right=s,top=s,bottom=s)

CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def hdr(ws, row, col, val, fill=HDR_FILL, font=WHT, align=CTR, span=1):
    c = ws.cell(row=row, column=col, value=val)
    c.fill=fill; c.font=font; c.alignment=align; c.border=bd()
    if span > 1:
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    return c

def cell(ws, row, col, val, fill=None, font=NRM, align=CTR, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill=fill
    c.font=font; c.alignment=align; c.border=bd()
    if fmt: c.number_format=fmt
    return c

# ── BUILD NEW SHEET ───────────────────────────────────────────────
def build_mes_400_sheet(ws):
    ws.sheet_view.showGridLines = False

    # Col widths
    widths = [8,6,7,7,7,7,7,7,7,7,7,7,10,10,10,10,10,10,10,10]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 14

    # ── TITLE ROW ─────────────────────────────────────────────────
    hdr(ws,1,1,
        'MES ATM Grid $400  |  Template: 3-5-75-1 heavyT1  |  '
        'T1=50%SL  T2=75%SL  T3=100%SL  |  '
        'BE trigger=T1, buffer=+2tk (+0.50pt)  |  '
        'Trail ALL contracts 8tk/6tk/4tk freq 2/2/1  |  $5/pt  $1.25/tick  fee=$0.57/contract',
        span=20)
    hdr(ws,2,1,
        'T1 qty = ceiling(total/2)  |  T2 = ceiling(rest/2)  |  T3 = remainder  |  '
        'SOP per SL: SL4-6=SOP3  SL7=SOP35  SL8=SOP4  SL9=SOP45  SL10=SOP5',
        fill=HDR2_FILL, span=20)

    # ── SECTION 1: ATM CONFIG ─────────────────────────────────────
    r = 4
    hdr(ws,r,1,'▶  ATM CONFIG  —  SL4 through SL10', fill=HDR3_FILL, span=20)
    r+=1
    cols = ['SL\n(pts)','SL\n(ticks)','Max\nRisk','Total\nQty','T1\nQty','T2\nQty','T3\nQty',
            'T1\n(pts)','T1\n(ticks)','T2\n(pts)','T2\n(ticks)','T3\n(pts)','T3\n(ticks)',
            'BE\nTrigger','BE\nStop','SOP','Trail\nStep1','Trail\nStep2','Trail\nStep3','Fees']
    for ci,col in enumerate(cols,1):
        hdr(ws,r,ci,col,fill=HDR2_FILL)
    r+=1

    for d in data:
        sl=d['sl']; steps=d['steps']; stops=d['stops']
        fill = ROW_FILL if sl%2==0 else ROW2_FILL
        vals = [
            sl, sl*4, f"${sl*5*d['total']:.0f}", d['total'],
            d['q1'], d['q2'], d['q3'],
            f"{d['t1v']}pt", ticks(d['t1v']),
            f"{d['t2v']}pt", ticks(d['t2v']),
            f"{d['t3v']}pt", ticks(d['t3v']),
            f"{d['t1v']}pt", '+0.50pt', d['sop'],
            f"{steps[0][0]}pt→{stops[0]}pt",
            f"{steps[1][0]}pt→{stops[1]}pt",
            f"{steps[2][0]}pt→{stops[2]}pt",
            f"${d['fee']}"
        ]
        for ci,v in enumerate(vals,1):
            cell(ws,r,ci,v,fill=fill)
        r+=1

    # ── SECTION 2: SCENARIO PnL ───────────────────────────────────
    r+=1
    hdr(ws,r,1,'▶  SCENARIO PnL  —  All outcomes per SL', fill=HDR3_FILL, span=20)
    r+=1
    scen_cols = ['SL','SOP','Full\nWin','Trail1\nstop','Trail2\nstop','Trail3\nstop',
                 'Auto BE\n(T1+2tk)','Hard\nStop','Quick\nBtn','BE Btn\n(EV 40/60)','E/trade']
    for ci,col in enumerate(scen_cols,1):
        hdr(ws,r,ci,col,fill=HDR2_FILL)
    r+=1

    for d in data:
        fill = ROW_FILL if d['sl']%2==0 else ROW2_FILL
        e_fill = GRN_FILL if d['e']>0 else RED_FILL
        vals = [
            f"SL{d['sl']}", d['sop'],
            f"+${d['full']:.2f}", f"+${d['tr1']:.2f}", f"+${d['tr2']:.2f}", f"+${d['tr3']:.2f}",
            f"+${d['auto_be']:.2f}", f"${d['hard']:.2f}",
            f"+${d['quick']:.2f}", f"${d['be_btn']:.2f}",
            f"${d['e']:+.2f}"
        ]
        fills = [fill,fill,GRN_FILL,GRN_FILL,GRN_FILL,GRN_FILL,GRN_FILL,RED_FILL,GRN_FILL,RED_FILL,e_fill]
        for ci,(v,f_) in enumerate(zip(vals,fills),1):
            cell(ws,r,ci,v,fill=f_)
        r+=1

    # ── SECTION 3: EXPECTANCY BREAKDOWN ──────────────────────────
    r+=1
    hdr(ws,r,1,
        '▶  EXPECTANCY BREAKDOWN  —  Probs: Full=20% Trail1=10% Trail2=10% Trail3=10% '
        'AutoBE=10% HardStop=10% Quick=10% BEbtn=20%',
        fill=HDR3_FILL, span=20)
    r+=1
    exp_cols = ['SL','Full\n20%','Trail1\n10%','Trail2\n10%','Trail3\n10%',
                'AutoBE\n10%','HardStop\n10%','Quick\n10%','BEbtn\n20%','E/trade']
    for ci,col in enumerate(exp_cols,1):
        hdr(ws,r,ci,col,fill=HDR2_FILL)
    r+=1

    sys_e=0
    for d in data:
        fill = ROW_FILL if d['sl']%2==0 else ROW2_FILL
        c_full   = round(probs['full']*d['full'],2)
        c_tr1    = round(probs['trail1']*d['tr1'],2)
        c_tr2    = round(probs['trail2']*d['tr2'],2)
        c_tr3    = round(probs['trail3']*d['tr3'],2)
        c_abe    = round(probs['auto_be']*d['auto_be'],2)
        c_stop   = round(probs['stop']*d['hard'],2)
        c_quick  = round(probs['quick']*d['quick'],2)
        c_be     = round(probs['be_btn']*d['be_btn'],2)
        sys_e   += weights.get(d['sl'],0)*d['e']
        e_fill   = GRN_FILL if d['e']>0 else RED_FILL
        vals = [f"SL{d['sl']}",
                f"+${c_full:.2f}",f"+${c_tr1:.2f}",f"+${c_tr2:.2f}",f"+${c_tr3:.2f}",
                f"+${c_abe:.2f}",f"${c_stop:.2f}",f"+${c_quick:.2f}",f"${c_be:.2f}",
                f"${d['e']:+.2f}"]
        fills=[fill,GRN_FILL,GRN_FILL,GRN_FILL,GRN_FILL,GRN_FILL,RED_FILL,GRN_FILL,RED_FILL,e_fill]
        for ci,(v,f_) in enumerate(zip(vals,fills),1):
            cell(ws,r,ci,v,fill=f_)
        r+=1

    # System weighted row
    hdr(ws,r,1,'SYSTEM WEIGHTED E/trade', fill=HDR2_FILL, span=9)
    c=ws.cell(row=r,column=10,value=f"${sys_e:+.2f}")
    c.fill = GRN_FILL if sys_e>0 else RED_FILL
    c.font=Font(bold=True,size=11)
    c.alignment=CTR; c.border=bd()
    hdr(ws,r,11,'SL4=45% SL5=30% SL6=15% SL7=7% SL8=3%',fill=SOP_FILL,
        font=Font(italic=True,size=9),span=10)
    r+=2

    # ── SECTION 4: SOP REFERENCE ──────────────────────────────────
    hdr(ws,r,1,'▶  STOP STRATEGY REFERENCE  —  5 SOPs  (trail ALL contracts, 8tk/6tk/4tk, freq 2/2/1)',
        fill=HDR3_FILL, span=20)
    r+=1
    sop_cols=['SOP','Used by','Trail trigger\nStep 1 (8tk)','Stop after\nStep 1',
              'Trail trigger\nStep 2 (6tk)','Stop after\nStep 2',
              'Trail trigger\nStep 3 (4tk)','Stop after\nStep 3','Notes']
    for ci,col in enumerate(sop_cols,1):
        hdr(ws,r,ci,col,fill=HDR2_FILL)
    r+=1

    sop_notes = {
        'SOP3':  'SL4/5/6 — T1 fills before trail fires',
        'SOP35': 'SL7 — trail fires at T1 level (half SL=3.5pt)',
        'SOP4':  'SL8 — trail fires at T1 level (half SL=4pt)',
        'SOP45': 'SL9 — trail fires at T1 level (half SL=4.5pt)',
        'SOP5':  'SL10 — trail fires at T1 level (half SL=5pt)',
    }
    sop_used = {
        'SOP3':'SL4, SL5, SL6','SOP35':'SL7','SOP4':'SL8','SOP45':'SL9','SOP5':'SL10'
    }
    for i,(sop_name,steps) in enumerate(trail_sops.items()):
        stops=[round(s[0]-s[1],2) for s in steps]
        fill = ROW_FILL if i%2==0 else ROW2_FILL
        vals=[sop_name, sop_used[sop_name],
              f"{steps[0][0]}pt ({ticks(steps[0][0])}tk)", f"{stops[0]}pt",
              f"{steps[1][0]}pt ({ticks(steps[1][0])}tk)", f"{stops[1]}pt",
              f"{steps[2][0]}pt ({ticks(steps[2][0])}tk)", f"{stops[2]}pt",
              sop_notes[sop_name]]
        for ci,v in enumerate(vals,1):
            cell(ws,r,ci,v,fill=fill)
        r+=1

    # ── SECTION 5: NAMING CONVENTION ─────────────────────────────
    r+=1
    hdr(ws,r,1,'▶  TEMPLATE NAMING CONVENTION', fill=HDR3_FILL, span=20)
    r+=1
    notes=[
        'Format: [targets]-[T1%SL]-[T2%SL]-([T3%SL]-)[heavy]',
        'This config: 3-5-75-1-1  (3 targets, T1=50%SL, T2=75%SL, T3=100%SL, heavy T1)',
        '%SL codes: 5=50%SL   75=75%SL   1=100%SL',
        'heavy: 0=equal thirds   1=heavy T1 (ceiling split)   2=heavy T2   3=heavy T3',
        'SOP suffix: SOP3=trail@3/4/5   SOP35=trail@3.5/4.5/5.5   SOP4=trail@4/5/6   etc.',
        'BE: trigger=T1 price, buffer=+2 ticks (+0.50pt above entry)',
        'Quick button: q1@1pt, q2@2pt, q3@2pt (same quantities)',
        'BE button: 40% scratch at entry (-fees only), 60% full hard stop',
    ]
    for note in notes:
        c=ws.cell(row=r,column=1,value=note)
        c.font=SML; c.alignment=LFT; c.fill=SOP_FILL
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=20)
        r+=1

    # freeze top 3 rows
    ws.freeze_panes='A4'

# ── LOAD v14, REPLACE TAB, SAVE AS v16 ───────────────────────────
wb = load_workbook(SRC)

# Remove the old tab
if TAB_TO_REPLACE in wb.sheetnames:
    del wb[TAB_TO_REPLACE]

# Insert new tab at same position (was index 1)
wb.create_sheet(TAB_TO_REPLACE, 1)
ws_new = wb[TAB_TO_REPLACE]
build_mes_400_sheet(ws_new)

wb.save(DEST)
print(f'Saved: {DEST}')
print(f'Sheets: {wb.sheetnames}')
